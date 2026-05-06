"""序列化辅助：Excel / GraphML / Markdown 格式化，供 export.py 调用。"""
from __future__ import annotations
import io
import xml.etree.ElementTree as ET


def build_excel(docs: list[dict], sections: list[dict], refs: list[dict], entities: list[dict]) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(fill_type="solid", fgColor="1B6BB5")

    def _hdr(ws, cols: list[str]) -> None:
        ws.append(cols)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

    wb = openpyxl.Workbook()

    # Sheet1 文档概览
    ws1 = wb.active
    ws1.title = "文档概览"
    _hdr(ws1, ["规范编号", "标题", "版本", "章节数", "图片数", "发布日期"])
    for d in docs:
        ws1.append([d.get("doc_id"), d.get("title"), d.get("version"),
                    d.get("section_count", 0), d.get("image_count", 0), d.get("issue_date")])

    # Sheet2 章节详情
    ws2 = wb.create_sheet("章节详情")
    _hdr(ws2, ["规范编号", "章节号", "标题", "层级", "内容摘要"])
    for s in sections:
        content = (s.get("content") or "")[:200]
        ws2.append([s.get("doc_id"), s.get("number"), s.get("title"), s.get("level"), content])

    # Sheet3 引用关系
    ws3 = wb.create_sheet("引用关系")
    _hdr(ws3, ["源文档", "目标文档"])
    for r in refs:
        ws3.append([r.get("source"), r.get("target")])

    # Sheet4 实体统计
    ws4 = wb.create_sheet("实体统计")
    _hdr(ws4, ["实体名称", "类型", "出现次数", "相关文档"])
    for e in entities:
        ws4.append([e.get("name"), e.get("type"), e.get("count"), e.get("docs")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_graphml(nodes: list[dict], edges: list[dict]) -> str:
    root = ET.Element("graphml", xmlns="http://graphml.graphdrawing.org/graphml")
    for key_id, attr_name in [("label", "label"), ("title", "title"), ("type", "type")]:
        k = ET.SubElement(root, "key", id=key_id, **{"for": "node", "attr.name": attr_name, "attr.type": "string"})
        _ = k  # suppress unused warning
    graph = ET.SubElement(root, "graph", id="G", edgedefault="directed")
    for n in nodes:
        node = ET.SubElement(graph, "node", id=str(n["id"]))
        for key, val in [("label", n.get("label", "")), ("title", n.get("title", "")), ("type", n.get("type", ""))]:
            d = ET.SubElement(node, "data", key=key)
            d.text = str(val) if val else ""
    for i, e in enumerate(edges):
        edge = ET.SubElement(graph, "edge", id=f"e{i}", source=str(e["source"]), target=str(e["target"]))
        d = ET.SubElement(edge, "data", key="label")
        d.text = e.get("label", "")
    return '<?xml version="1.0" encoding="UTF-8"?>\n' + ET.tostring(root, encoding="unicode")


def build_markdown(docs: list[dict], sections: list[dict]) -> str:
    lines: list[str] = []
    sec_by_doc: dict[str, list[dict]] = {}
    for s in sections:
        sec_by_doc.setdefault(s["doc_id"], []).append(s)
    for d in docs:
        lines.append(f"# {d.get('doc_id')} {d.get('title', '')}\n")
        if d.get("version"):
            lines.append(f"**版本**: {d['version']}  **发布日期**: {d.get('issue_date', '')}\n\n")
        for s in sec_by_doc.get(d["doc_id"], []):
            depth = min(s.get("level", 1), 4)
            lines.append(f"{'#' * (depth + 1)} {s.get('number', '')} {s.get('title', '')}\n\n")
            if s.get("content"):
                lines.append(f"{s['content']}\n\n")
    return "".join(lines)
